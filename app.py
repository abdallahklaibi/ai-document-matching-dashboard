import io
import re
import unicodedata
from datetime import datetime
from difflib import SequenceMatcher

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

try:
    import fitz  # PyMuPDF
    import pytesseract
    from PIL import Image, ImageOps, ImageEnhance
except Exception:
    fitz = None
    pytesseract = None
    Image = None
    ImageOps = None
    ImageEnhance = None

st.set_page_config(page_title="Customer Record Reconciliation", page_icon="📊", layout="wide")

st.markdown("""
<style>
.block-container {padding-top: 2rem; padding-bottom: 3rem; max-width: 1450px;}
.hero {padding: 24px 28px; border: 1px solid rgba(120,120,120,.20); border-radius: 18px; margin-bottom: 22px;}
.hero h1 {margin:0; font-size:2rem;}
.hero p {margin:.5rem 0 0 0; opacity:.75;}
.step {border: 1px solid rgba(120,120,120,.22); border-radius: 14px; padding: 14px 16px; min-height: 85px;}
.small-muted {opacity:.67; font-size:.9rem;}
div[data-testid="stMetric"] {border: 1px solid rgba(120,120,120,.20); padding: 16px; border-radius: 14px;}
</style>
""", unsafe_allow_html=True)

# -------------------- Normalisation --------------------

def clean_text(value):
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def normalize_latin_name(value):
    s = clean_text(value)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.upper()
    s = re.sub(r"[^A-Z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def digits_only(value):
    return re.sub(r"\D", "", clean_text(value))


MOBILE_PREFIXES = {"050", "052", "054", "055", "056", "058"}
LANDLINE_PREFIXES = {"02", "03", "04", "06", "07", "09"}


def normalize_phone(value):
    d = digits_only(value)
    if not d:
        return ""
    if d.startswith("00971"):
        d = d[2:]
    if d.startswith("971"):
        n = d[3:]
        if len(n) == 9 and ("0" + n[:2]) in MOBILE_PREFIXES:
            return "0" + n
        if len(n) == 8 and ("0" + n[:1]) in LANDLINE_PREFIXES:
            return "0" + n
        return ""
    if len(d) == 10 and d[:3] in MOBILE_PREFIXES:
        return d
    if len(d) == 9 and d[:2] in LANDLINE_PREFIXES:
        return d
    if len(d) == 9 and ("0" + d[:2]) in MOBILE_PREFIXES:
        return "0" + d
    if len(d) == 8 and ("0" + d[:1]) in LANDLINE_PREFIXES:
        return "0" + d
    return ""


def extract_uae_phones(text):
    text = clean_text(text)
    patterns = [
        r"(?<!\d)(?:\+?971|00971)[\s()./-]*5[024568](?:[\s()./-]*\d){7}(?!\d)",
        r"(?<!\d)0?5[024568](?:[\s()./-]*\d){7}(?!\d)",
        r"(?<!\d)(?:\+?971|00971)[\s()./-]*[234679](?:[\s()./-]*\d){7}(?!\d)",
        r"(?<!\d)0[234679](?:[\s()./-]*\d){7}(?!\d)",
    ]
    found, seen = [], set()
    for pattern in patterns:
        for m in re.finditer(pattern, text):
            p = normalize_phone(m.group(0))
            if p and p not in seen:
                seen.add(p)
                found.append(p)
    return found


def normalize_account(value):
    d = digits_only(value)
    if 4 <= len(d) <= 8:
        return d.lstrip("0") or "0"
    return ""


def extract_account_candidates(text):
    """Prefer short numeric values written in brackets; then other 4-8 digit values."""
    text = clean_text(text)
    phones = set(extract_uae_phones(text))
    phone_digit_strings = {digits_only(p) for p in phones}
    out, seen = [], set()

    # Parenthesized numbers are especially useful on these customer forms.
    patterns = [r"[\(\[]\s*([0-9][0-9\s-]{2,9}[0-9])\s*[\)\]]", r"(?<!\d)(\d{4,8})(?!\d)"]
    for pat in patterns:
        for m in re.finditer(pat, text):
            d = digits_only(m.group(1))
            if not (4 <= len(d) <= 8):
                continue
            # Avoid obvious years/dates/times and chunks already part of a phone.
            if d in {"2026", "2025", "2024"}:
                continue
            if any(d in ph for ph in phone_digit_strings):
                continue
            norm = normalize_account(d)
            if norm and norm not in seen:
                seen.add(norm)
                out.append(norm)
    return out


# -------------------- Excel indexes --------------------

def choose_column(df, preferred_names, fallback=None):
    lookup = {str(c).strip().lower(): c for c in df.columns}
    for name in preferred_names:
        if name.lower() in lookup:
            return lookup[name.lower()]
    return fallback if fallback in df.columns else (df.columns[0] if len(df.columns) else None)


def build_indexes(df, name_col, account_col, phone_col):
    account_map = {}
    phone_map = {}
    name_rows = []

    for idx, row in df.iterrows():
        account = normalize_account(row.get(account_col, "")) if account_col else ""
        if account:
            account_map.setdefault(account, []).append(idx)

        if phone_col:
            raw = row.get(phone_col, "")
            phones = extract_uae_phones(raw)
            if not phones:
                p = normalize_phone(raw)
                phones = [p] if p else []
            for p in phones:
                phone_map.setdefault(p, []).append(idx)

        if name_col:
            raw_name = clean_text(row.get(name_col, ""))
            name_norm = normalize_latin_name(raw_name)
            if len(name_norm) >= 3:
                name_rows.append((idx, raw_name, name_norm))

    return account_map, phone_map, name_rows


# -------------------- OCR --------------------

def preprocess_image(image):
    gray = ImageOps.grayscale(image)
    gray = ImageOps.autocontrast(gray)
    gray = ImageEnhance.Contrast(gray).enhance(1.35)
    return gray


def ocr_top_form(image):
    """Fast OCR: read only the upper customer-information section of the form."""
    w, h = image.size
    # The useful customer fields are in the upper part of this form.
    crop = image.crop((int(w * 0.07), int(h * 0.06), int(w * 0.84), int(h * 0.34)))
    crop = preprocess_image(crop)
    # One sparse-text pass is much faster than OCR-ing the entire page twice.
    return pytesseract.image_to_string(crop, config="--psm 11").strip()


def render_page(page, dpi=175):
    matrix = fitz.Matrix(dpi / 72.0, dpi / 72.0)
    pix = page.get_pixmap(matrix=matrix, alpha=False)
    return Image.open(io.BytesIO(pix.tobytes("png")))


# -------------------- Name matching --------------------

def name_similarity(name_norm, ocr_norm):
    if not name_norm or not ocr_norm:
        return 0.0
    # Fast exact containment is strong evidence.
    if name_norm in ocr_norm:
        return 1.0
    name_tokens = name_norm.split()
    ocr_tokens = ocr_norm.split()
    if not name_tokens or not ocr_tokens:
        return 0.0

    n = len(name_tokens)
    best = 0.0
    # Compare against sliding OCR windows around the same token length.
    for size in range(max(1, n - 1), min(len(ocr_tokens), n + 2) + 1):
        for i in range(0, len(ocr_tokens) - size + 1):
            window = " ".join(ocr_tokens[i:i + size])
            score = SequenceMatcher(None, name_norm, window).ratio()
            if score > best:
                best = score
    # Also reward token overlap (helps swapped first/last names).
    name_set = set(name_tokens)
    ocr_set = set(ocr_tokens)
    token_overlap = len(name_set & ocr_set) / max(len(name_set), 1)
    return max(best, token_overlap)


def best_name_matches(ocr_text, name_rows, limit=3):
    ocr_norm = normalize_latin_name(ocr_text)
    scored = []
    for idx, raw_name, name_norm in name_rows:
        score = name_similarity(name_norm, ocr_norm)
        if score >= 0.45:
            scored.append((score, idx, raw_name))
    scored.sort(reverse=True, key=lambda x: x[0])
    return scored[:limit]


# -------------------- Evidence fusion --------------------

def evaluate_page(ocr_text, page_no, df, name_col, account_col, phone_col, account_map, phone_map, name_rows):
    """
    Smart fallback order requested by the user:
    1) Phone first.
    2) If phone is unclear / not found in Excel, try Name.
    3) If name is unclear / ambiguous, try Account Number.

    Agreement from a second field upgrades confidence when available.
    """
    phones = extract_uae_phones(ocr_text)
    accounts = extract_account_candidates(ocr_text)
    name_candidates = best_name_matches(ocr_text, name_rows, limit=5)

    chosen_idx = None
    method = ""
    confidence = "Not Matched"
    score = 0
    evidence = []
    matched = False

    # ---------- 1. PHONE FIRST ----------
    phone_hits = []
    for ph in phones:
        for idx in phone_map.get(ph, []):
            phone_hits.append((idx, ph))

    unique_phone_rows = sorted(set(idx for idx, _ in phone_hits))
    if len(unique_phone_rows) == 1:
        chosen_idx = unique_phone_rows[0]
        matched_phone = next(ph for idx, ph in phone_hits if idx == chosen_idx)
        method = "Phone"
        evidence.append(f"Phone {matched_phone}")
        score = 92
        confidence = "High"
        matched = True

        # Optional supporting evidence can only strengthen the result.
        row_account = normalize_account(df.loc[chosen_idx, account_col]) if account_col else ""
        if row_account and row_account in accounts:
            evidence.append(f"Account {row_account}")
            score = 100
        for sim, idx, raw_name in name_candidates:
            if idx == chosen_idx and sim >= 0.72:
                evidence.append(f"Name {sim:.0%}")
                score = min(100, score + 5)
                break

    # ---------- 2. NAME FALLBACK ----------
    if chosen_idx is None and name_candidates:
        best_sim, best_idx, raw_name = name_candidates[0]
        second_sim = name_candidates[1][0] if len(name_candidates) > 1 else 0.0
        margin = best_sim - second_sim

        # A clear fuzzy-name winner is enough to proceed.
        if best_sim >= 0.80 and margin >= 0.08:
            chosen_idx = best_idx
            method = "Name"
            evidence.append(f"Name {best_sim:.0%}")
            score = int(min(90, 60 + best_sim * 30))
            confidence = "Medium"
            matched = True

            # Exact account/phone agreement upgrades to High.
            row_account = normalize_account(df.loc[chosen_idx, account_col]) if account_col else ""
            row_phones = extract_uae_phones(df.loc[chosen_idx, phone_col]) if phone_col else []
            if row_account and row_account in accounts:
                evidence.append(f"Account {row_account}")
                score = min(100, score + 15)
                confidence = "High"
            if set(row_phones) & set(phones):
                common = list(set(row_phones) & set(phones))[0]
                evidence.append(f"Phone {common}")
                score = 100
                confidence = "High"
        elif best_sim >= 0.65 and margin >= 0.05:
            # Keep a plausible but weak name for manual review; account fallback may still rescue it below.
            chosen_idx = best_idx
            method = "Name"
            evidence.append(f"Name {best_sim:.0%}")
            score = int(45 + best_sim * 20)
            confidence = "Manual Review"
            matched = False

    # ---------- 3. ACCOUNT FALLBACK ----------
    # Use account number when no automatic phone/name match was achieved.
    if not matched:
        account_hits = []
        for acc in accounts:
            for idx in account_map.get(acc, []):
                account_hits.append((idx, acc))
        unique_account_rows = sorted(set(idx for idx, _ in account_hits))

        if len(unique_account_rows) == 1:
            acc_idx = unique_account_rows[0]
            matched_acc = next(acc for idx, acc in account_hits if idx == acc_idx)
            chosen_idx = acc_idx
            method = "Account Number"
            evidence = [f"Account {matched_acc}"]
            score = 88
            confidence = "High"
            matched = True

            # Supporting name evidence upgrades/validates.
            for sim, idx, raw_name in name_candidates:
                if idx == chosen_idx and sim >= 0.65:
                    evidence.append(f"Name {sim:.0%}")
                    score = min(100, score + 8)
                    break
        elif len(unique_account_rows) > 1 and chosen_idx is None:
            # Account exists but is not unique; do not auto-match it.
            chosen_idx = unique_account_rows[0]
            method = "Account Number"
            evidence = ["Account number is not unique in Excel"]
            score = 45
            confidence = "Manual Review"
            matched = False

    if chosen_idx is None:
        return {
            "Page": page_no,
            "Match Method": "No clear field",
            "OCR Name/Text": " ".join(normalize_latin_name(ocr_text).split()[:18]),
            "PDF Account": accounts[0] if accounts else "",
            "PDF Phone": phones[0] if phones else "",
            "Matched": False,
            "Confidence": "Not Matched",
            "Match Score": 0,
            "Excel Row": None,
            "Excel Name": "",
            "Excel Account": "",
            "Excel Phone": "",
            "Evidence": "Phone unclear → Name unclear → Account unclear",
        }

    excel_name = clean_text(df.loc[chosen_idx, name_col]) if name_col else ""
    excel_account = clean_text(df.loc[chosen_idx, account_col]) if account_col else ""
    excel_phone = clean_text(df.loc[chosen_idx, phone_col]) if phone_col else ""

    return {
        "Page": page_no,
        "Match Method": method,
        "OCR Name/Text": " ".join(normalize_latin_name(ocr_text).split()[:18]),
        "PDF Account": accounts[0] if accounts else "",
        "PDF Phone": phones[0] if phones else "",
        "Matched": matched,
        "Confidence": confidence,
        "Match Score": int(score),
        "Excel Row": int(chosen_idx) + 2,
        "Excel Name": excel_name,
        "Excel Account": excel_account,
        "Excel Phone": excel_phone,
        "Evidence": ", ".join(evidence),
    }


# -------------------- Output --------------------

def create_output_excel(original_df, results_df):
    export_df = original_df.copy()
    for col, default in [
        ("Match_Status", "Not Matched"),
        ("Match_Confidence", ""),
        ("Source_Page", ""),
        ("Match_Score", ""),
        ("Match_Evidence", ""),
    ]:
        export_df[col] = pd.Series([default] * len(export_df), index=export_df.index, dtype="object")

    matched_rows = results_df[results_df["Matched"] == True].copy()
    if not matched_rows.empty:
        for _, r in matched_rows.sort_values("Match Score", ascending=False).iterrows():
            if pd.isna(r["Excel Row"]):
                continue
            idx = int(r["Excel Row"]) - 2
            if idx not in export_df.index:
                continue
            existing = export_df.loc[idx, "Match_Score"]
            existing_score = int(existing) if str(existing).isdigit() else -1
            if int(r["Match Score"]) >= existing_score:
                export_df.loc[idx, "Match_Status"] = "Matched"
                export_df.loc[idx, "Match_Confidence"] = r["Confidence"]
                export_df.loc[idx, "Source_Page"] = int(r["Page"])
                export_df.loc[idx, "Match_Score"] = int(r["Match Score"])
                export_df.loc[idx, "Match_Evidence"] = r["Evidence"]

    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Matched Data")
        results_df.to_excel(writer, index=False, sheet_name="Page Audit")

    bio.seek(0)
    wb = load_workbook(bio)
    green = PatternFill("solid", fgColor="E2F0D9")
    yellow = PatternFill("solid", fgColor="FFF2CC")
    orange = PatternFill("solid", fgColor="FCE4D6")
    gray = PatternFill("solid", fgColor="E7E6E6")

    ws = wb["Matched Data"]
    headers = {cell.value: cell.column for cell in ws[1]}
    status_col = headers.get("Match_Status")
    conf_col = headers.get("Match_Confidence")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")

    for r in range(2, ws.max_row + 1):
        status = ws.cell(r, status_col).value if status_col else ""
        conf = ws.cell(r, conf_col).value if conf_col else ""
        if status == "Matched" and conf == "High":
            fill = green
        elif status == "Matched" and conf == "Medium":
            fill = yellow
        elif conf == "Manual Review":
            fill = orange
        else:
            fill = gray
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).fill = fill

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for col_cells in sheet.columns:
            max_len = min(max(len(str(c.value or "")) for c in col_cells) + 2, 45)
            sheet.column_dimensions[col_cells[0].column_letter].width = max_len

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# -------------------- UI --------------------
st.markdown("""
<div class="hero">
  <h1>Customer Record Matching & Reconciliation</h1>
  <p>Free PDF-to-Excel matching with smart fallback: Phone → Name → Account Number, using local OCR and no paid AI service.</p>
</div>
""", unsafe_allow_html=True)

steps = st.columns(4)
for c, title, desc in zip(
    steps,
    ["1 · Upload", "2 · Read", "3 · Match", "4 · Export"],
    ["Scanned PDF + Excel", "Fast top-section OCR", "Phone → Name → Account", "Highlighted Excel"],
):
    c.markdown(f'<div class="step"><b>{title}</b><br><span class="small-muted">{desc}</span></div>', unsafe_allow_html=True)

st.write("")
left, right = st.columns(2)
with left:
    pdf_file = st.file_uploader("Monthly scanned PDF", type=["pdf"])
with right:
    excel_file = st.file_uploader("Monthly Excel", type=["xlsx", "xls"])

df = None
if excel_file is not None:
    try:
        df = pd.read_excel(excel_file)
        st.success(f"Excel loaded: {len(df):,} rows")
    except Exception as e:
        st.error(f"Could not read Excel: {e}")

name_col = account_col = phone_col = None
if df is not None and len(df.columns):
    columns = list(df.columns)
    default_name = choose_column(df, ["businessname", "woCompletedSiteContact", "wocompletedSiteBussinessName", "COMBOname"])
    default_account = choose_column(df, ["accountnum"])
    default_phone = choose_column(df, ["woCompletedWorkOrderSitePhone", "phone", "mobile"])

    c1, c2, c3 = st.columns(3)
    with c1:
        name_col = st.selectbox("Excel name column", columns, index=columns.index(default_name) if default_name in columns else 0)
    with c2:
        account_col = st.selectbox("Excel account-number column", columns, index=columns.index(default_account) if default_account in columns else 0)
    with c3:
        phone_col = st.selectbox("Excel phone column", columns, index=columns.index(default_phone) if default_phone in columns else 0)

    with st.expander("Preview selected Excel fields"):
        st.dataframe(df[[name_col, account_col, phone_col]].head(12), use_container_width=True, hide_index=True)

st.caption("No Claude, no API key and no paid AI service. OCR runs with open-source Tesseract on the Streamlit server.")

run_mode = st.radio(
    "Run mode",
    ["Quick Test — first 10 pages", "Full Monthly Run — all pages"],
    horizontal=True,
    help="Use Quick Test first to verify matching before processing the full PDF.",
)

run = st.button(
    "Run Monthly Analysis",
    type="primary",
    use_container_width=True,
    disabled=(pdf_file is None or df is None or name_col is None or account_col is None or phone_col is None),
)

if run:
    if fitz is None or pytesseract is None or Image is None:
        st.error("OCR components are not installed. Check requirements.txt and packages.txt.")
    else:
        progress = st.progress(0, text="Preparing matching indexes…")
        try:
            account_map, phone_map, name_rows = build_indexes(df, name_col, account_col, phone_col)
            pdf_bytes = pdf_file.getvalue()
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            pdf_total_pages = len(doc)
            total_pages = min(10, pdf_total_pages) if run_mode.startswith("Quick Test") else pdf_total_pages
            results = []

            for page_index in range(total_pages):
                page = doc.load_page(page_index)
                image = render_page(page, dpi=175)
                ocr_text = ocr_top_form(image)
                result = evaluate_page(
                    ocr_text, page_index + 1, df,
                    name_col, account_col, phone_col,
                    account_map, phone_map, name_rows,
                )
                results.append(result)
                pct = 5 + int((page_index + 1) / max(total_pages, 1) * 75)
                progress.progress(min(pct, 80), text=f"Reading and matching page {page_index + 1}/{total_pages}…")

            doc.close()
            results_df = pd.DataFrame(results)
            progress.progress(88, text="Preparing highlighted Excel report…")
            output_bytes = create_output_excel(df, results_df)
            progress.progress(100, text="Analysis complete")

            st.session_state["multi_results"] = results_df
            st.session_state["multi_output"] = output_bytes
        except Exception as e:
            progress.empty()
            st.error(f"Analysis failed: {e}")

if "multi_results" in st.session_state:
    results = st.session_state["multi_results"]
    total = len(results)
    matched = int(results["Matched"].sum()) if total else 0
    manual = int((results["Confidence"] == "Manual Review").sum()) if total else 0
    rate = (matched / total * 100) if total else 0

    st.divider()
    st.subheader("Executive Summary")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("PDF Pages Reviewed", f"{total:,}")
    c2.metric("Matched Pages", f"{matched:,}")
    c3.metric("Manual Review", f"{manual:,}")
    c4.metric("Match Rate", f"{rate:.1f}%")

    st.subheader("Page-by-Page Match Results")
    f1, f2 = st.columns([1, 2])
    with f1:
        conf_filter = st.multiselect(
            "Confidence",
            ["High", "Medium", "Manual Review", "Not Matched"],
            default=["High", "Medium", "Manual Review", "Not Matched"],
        )
    with f2:
        search = st.text_input("Search name / account / phone / page")

    filtered = results[results["Confidence"].isin(conf_filter)].copy()
    if search:
        mask = filtered.astype(str).apply(lambda col: col.str.contains(search, case=False, na=False)).any(axis=1)
        filtered = filtered[mask]

    st.dataframe(
        filtered,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Match Score": st.column_config.ProgressColumn("Match Score", min_value=0, max_value=100, format="%d%%"),
        },
    )

    stamp = datetime.now().strftime("%Y-%m-%d")
    st.download_button(
        "Download Highlighted Excel Report",
        data=st.session_state["multi_output"],
        file_name=f"Customer_Matched_Report_{stamp}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )

    with st.expander("How matching works"):
        st.markdown("""
- The app follows the requested fallback order: **Phone → Name → Account Number**.
- A clear **phone** match is accepted first. If the phone is unclear or does not match Excel, the app tries the **name**.
- If the name is also unclear or ambiguous, it tries the **account number**.
- A second agreeing field can upgrade the confidence.
- **High** means strong evidence with a clear winner.
- **Medium** means a likely match that should still be spot-checked.
- **Manual Review** means the page has some evidence, but not enough for automatic acceptance.
""")

st.divider()
st.caption("Free local OCR supports reconciliation; final business decisions remain with the responsible reviewer.")
